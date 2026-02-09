"""
Tests for Spreadsheet Creation
=================================
Tests CSV and Excel file creation.
"""

import csv

import pytest

from kai.io.spreadsheet import create_csv, create_excel


@pytest.fixture
def sample_data():
    """Sample spreadsheet data."""
    return {
        "headers": ["Name", "Age", "City"],
        "rows": [
            ["Alice", 30, "New York"],
            ["Bob", 25, "Los Angeles"],
            ["Charlie", 35, "Chicago"],
        ],
    }


def test_create_csv(tmp_path, sample_data):
    """Test CSV creation with headers and rows."""
    path = tmp_path / "test.csv"
    result = create_csv(str(path), sample_data["headers"], sample_data["rows"])

    assert result.exists()
    assert result.suffix == ".csv"

    # Read back and verify
    with open(result, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        rows = list(reader)

    assert rows[0] == ["Name", "Age", "City"]
    assert rows[1] == ["Alice", "30", "New York"]  # CSV converts to strings
    assert len(rows) == 4  # header + 3 data rows


def test_create_csv_empty_rows(tmp_path):
    """Test CSV with headers only."""
    path = tmp_path / "empty.csv"
    result = create_csv(str(path), ["A", "B"], [])

    assert result.exists()
    with open(result, encoding="utf-8") as f:
        content = f.read()
    assert "A,B" in content


def test_create_csv_creates_parent_dirs(tmp_path):
    """Test that CSV creation creates parent directories."""
    path = tmp_path / "deep" / "nested" / "data.csv"
    result = create_csv(str(path), ["X"], [[1], [2]])
    assert result.exists()


def test_create_excel(tmp_path, sample_data):
    """Test Excel creation with headers and rows."""
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not installed")

    path = tmp_path / "test.xlsx"
    result = create_excel(str(path), sample_data["headers"], sample_data["rows"])

    assert result.exists()
    assert result.suffix == ".xlsx"

    # Read back and verify
    wb = openpyxl.load_workbook(result)
    ws = wb.active
    assert ws.title == "Sheet1"
    assert ws.cell(row=1, column=1).value == "Name"
    assert ws.cell(row=2, column=1).value == "Alice"
    assert ws.cell(row=2, column=2).value == 30  # Preserves types
    assert ws.max_row == 4  # header + 3 rows


def test_create_excel_custom_sheet_name(tmp_path):
    """Test Excel with custom sheet name."""
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not installed")

    path = tmp_path / "custom.xlsx"
    create_excel(str(path), ["Col"], [[1]], sheet_name="MyData")

    wb = openpyxl.load_workbook(path)
    assert wb.active.title == "MyData"


def test_create_excel_creates_parent_dirs(tmp_path):
    """Test that Excel creation creates parent directories."""
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not installed")

    path = tmp_path / "deep" / "nested" / "data.xlsx"
    result = create_excel(str(path), ["X"], [[1]])
    assert result.exists()
