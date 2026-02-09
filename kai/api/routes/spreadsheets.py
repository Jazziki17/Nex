"""Spreadsheet creation route — Excel and CSV."""

import csv
import io
from pathlib import Path

from fastapi import APIRouter, HTTPException
from pydantic import BaseModel

from kai.utils.logger import setup_logger

logger = setup_logger(__name__)

router = APIRouter(tags=["spreadsheets"])


class SpreadsheetRequest(BaseModel):
    path: str
    headers: list[str]
    rows: list[list[str | int | float | None]]
    format: str = "xlsx"  # "xlsx" or "csv"


def _safe_path(path_str: str) -> Path:
    """Resolve and validate path is under home directory."""
    resolved = Path(path_str).expanduser().resolve()
    home = Path.home().resolve()
    if not str(resolved).startswith(str(home)):
        raise ValueError(f"Path not within home directory: {path_str}")
    return resolved


@router.post("/spreadsheet")
async def create_spreadsheet(req: SpreadsheetRequest):
    """Create an Excel or CSV file from structured data."""
    try:
        file_path = _safe_path(req.path)
    except ValueError as e:
        raise HTTPException(status_code=403, detail=str(e))

    file_path.parent.mkdir(parents=True, exist_ok=True)

    if req.format == "csv":
        return _write_csv(file_path, req.headers, req.rows)
    elif req.format == "xlsx":
        return _write_xlsx(file_path, req.headers, req.rows)
    else:
        raise HTTPException(status_code=400, detail=f"Unsupported format: {req.format}")


def _write_csv(path: Path, headers: list[str], rows: list[list]) -> dict:
    """Write data as CSV."""
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)

    logger.info(f"CSV created: {path}")
    return {"path": str(path), "format": "csv", "rows": len(rows), "status": "ok"}


def _write_xlsx(path: Path, headers: list[str], rows: list[list]) -> dict:
    """Write data as Excel (requires openpyxl)."""
    try:
        from openpyxl import Workbook
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl not installed. Run: pip install openpyxl",
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Write headers
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Write data rows
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(path)
    logger.info(f"Excel created: {path}")
    return {"path": str(path), "format": "xlsx", "rows": len(rows), "status": "ok"}
