"""
Spreadsheet Creation Module
=============================
Create Excel (.xlsx) and CSV files from structured data.
"""

import csv
from pathlib import Path

from nex.utils.logger import setup_logger

logger = setup_logger(__name__)


def create_csv(path: str | Path, headers: list[str], rows: list[list]) -> Path:
    """Create a CSV file from headers and rows."""
    path = Path(path).expanduser().resolve()
    path.parent.mkdir(parents=True, exist_ok=True)

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)

    logger.info(f"CSV created: {path} ({len(rows)} rows)")
    return path


def create_excel(path: str | Path, headers: list[str], rows: list[list],
                 sheet_name: str = "Sheet1") -> Path:
    """Create an Excel file from headers and rows. Requires openpyxl."""
    from openpyxl import Workbook

    path = Path(path).expanduser().resolve()
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(path)
    logger.info(f"Excel created: {path} ({len(rows)} rows)")
    return path
